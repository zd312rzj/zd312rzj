import os
import zipfile
import shutil
import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as xlImage
import time
import re

# ================= 配置参数 =================
# txt 文件的绝对路径（注意转义反斜杠）
CATEGORY_FILE_PATH = r"C:\Users\zhangjian08\Desktop\二级类目.txt" # 这里请替换成你实际的 txt 文件路径
MONTH_STR = "202605"  # 账期字符串
DESKTOP_PATH = os.path.join(os.path.expanduser("~"), "Desktop", "账单文件") # 最终文件保存的桌面文件夹路径

def load_categories():
    """从 txt 文件中读取业务大类列表"""
    categories = []
    try:
        # 使用 utf-8 编码读取，如果报错会尝试 gbk
        with open(CATEGORY_FILE_PATH, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                # 去除每行前后的空格和换行符
                cat = line.strip()
                # 忽略空行
                if cat:
                    categories.append(cat)
    except UnicodeDecodeError:
        # Windows 的记事本默认可能是 ANSI (GBK) 编码
        with open(CATEGORY_FILE_PATH, 'r', encoding='gbk') as f:
            lines = f.readlines()
            for line in lines:
                cat = line.strip()
                if cat:
                    categories.append(cat)
    except Exception as e:
        print(f"❌ 读取业务大类文件失败: {e}")
        print("请检查文件路径是否正确！")
        exit(1)
        
    return categories

# 动态加载所有的业务大类
BUSINESS_CATEGORIES = load_categories()

#================== 功能函数 =================
def setup_directories():
    if not os.path.exists(DESKTOP_PATH):
        os.makedirs(DESKTOP_PATH)

def process_excel_and_images(original_excel_path, amount, category, images):
    """处理Excel表格：插入截图并重命名（支持自动解压 ZIP 和 CSV 转换）"""
    new_filename = f"{amount}-{category}_{MONTH_STR}_{MONTH_STR}.xlsx"
    final_path = os.path.join(DESKTOP_PATH, new_filename)
    
    actual_file_to_read = original_excel_path
    
    # 1. 如果下载下来的是 ZIP 压缩包，先解压它
    if original_excel_path.endswith('.zip'):
        print(f"检测到 ZIP 压缩包，正在解压: {original_excel_path}")
        extract_dir = f"temp_extract_{category}"
        os.makedirs(extract_dir, exist_ok=True)
        with zipfile.ZipFile(original_excel_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
            extracted_files = os.listdir(extract_dir)
            if extracted_files:
                actual_file_to_read = os.path.join(extract_dir, extracted_files[0])
                print(f"解压得到明细文件: {actual_file_to_read}")

    # 2. 读取数据文件并保存为最终的 XLSX
    try:
        if actual_file_to_read.endswith('.csv'):
            df = pd.read_csv(actual_file_to_read, encoding='gbk', dtype=str) 
            df.to_excel(final_path, index=False)
        else:
            df = pd.read_excel(actual_file_to_read, dtype=str)
            df.to_excel(final_path, index=False)
    except Exception as e:
        print(f"⚠️ 默认编码读取表格失败，尝试使用 utf-8 编码读取 CSV... 错误: {e}")
        try:
            df = pd.read_csv(actual_file_to_read, encoding='utf-8', dtype=str)
            df.to_excel(final_path, index=False)
        except Exception as e2:
            print(f"❌ 表格转换彻底失败: {e2}")
            return 

    # --- 确保图片已经完全写入硬盘 ---
    for img_path in images.values():
        if os.path.exists(img_path):
            # 等待图片文件体积大于 0（最多等 2 秒）
            retries = 10
            while os.path.getsize(img_path) == 0 and retries > 0:
                time.sleep(0.2)
                retries -= 1

    # 3. 使用 openpyxl 插入图片
    wb = load_workbook(final_path)
    ws_bill = wb.create_sheet("账单截图")
    ws_invoice = wb.create_sheet("发票截图")
    
    # 插入账单截图
    if os.path.exists(images['bill']) and os.path.getsize(images['bill']) > 0:
        try:
            img = xlImage(images['bill'])
            ws_bill.add_image(img, 'A1')
            print("✅ 账单截图已插入 Excel")
        except Exception as e:
            print(f"⚠️ 插入账单截图失败: {e}")
            
    # 插入发票截图1和2
    if os.path.exists(images['invoice1']) and os.path.getsize(images['invoice1']) > 0:
        try:
            img1 = xlImage(images['invoice1'])
            ws_invoice.add_image(img1, 'A1')
            print("✅ 发票截图1已插入 Excel")
        except Exception as e:
            print(f"⚠️ 插入发票截图1失败: {e}")
            
    if os.path.exists(images['invoice2']) and os.path.getsize(images['invoice2']) > 0:
        try:
            img2 = xlImage(images['invoice2'])
            # 【核心修改】：将第二张图的起始位置设为 A60，预留足够空间
            ws_invoice.add_image(img2, 'A60') 
            print("✅ 发票截图2已插入 Excel")
        except Exception as e:
            print(f"⚠️ 插入发票截图2失败: {e}")
        
    wb.save(final_path)
    print(f"✅ 文件处理完成，已保存至桌面文件夹: {final_path}")
    
    # 清理 ZIP 解压产生的临时文件夹
    if original_excel_path.endswith('.zip'):
        shutil.rmtree(extract_dir, ignore_errors=True)

def main():
    setup_directories()
    
    with sync_playwright() as p:
        print("正在尝试接管已打开的 Edge 浏览器 (端口: 9222)...")
        browser = p.chromium.connect_over_cdp("http://127.0.0.1:9222")
        context = browser.contexts[0]
        
        # 寻找已打开的天猫页面
        page = None
        for p_item in context.pages:
            if "myseller.taobao.com" in p_item.url or "天猫" in p_item.title():
                page = p_item
                break
        
        if page is None:
            page = context.new_page()
            page.goto("https://myseller.taobao.com/")
        else:
            page.bring_to_front()
            
        print("✅ 成功接管浏览器！开始执行自动化任务...")

        # 【新增】：记录循环开始前，浏览器里原有的标签页数量
        base_pages_count = len(context.pages)
        
        for category in BUSINESS_CATEGORIES:
            print(f"\n================ 开始处理业务大类: {category} ================")
            images_paths = {
                'bill': f'temp_bill_{category}.png',
                'invoice1': f'temp_inv1_{category}.png',
                'invoice2': f'temp_inv2_{category}.png'
            }


            # 【核心修改】：每次循环开始前，强制新建一个干净的标签页，保证绝对安全
            try:
                page = context.new_page()
                page.bring_to_front()
                page.goto("https://myseller.taobao.com/", timeout=15000)
                page.wait_for_timeout(3000)
            except Exception as e:
                print(f"⚠️ 新建初始标签页失败，尝试使用现存页面: {e}")
                if len(context.pages) > 0:
                    page = context.pages[-1]
                    page.bring_to_front()
                    page.goto("https://myseller.taobao.com/")
                    page.wait_for_timeout(3000)
            
            # ================= 第一阶段: 账单查询与截图 =================
            try:
                page.get_by_text("财务", exact=True).first.click(force=True)
                page.wait_for_timeout(1500) 
                
                # 记录点击前的页面数量
                pages_count_before = len(context.pages)
                
                page.get_by_text("收支账单", exact=True).first.click(force=True)
                page.wait_for_timeout(3000)
                
                # 处理可能弹出的新标签页
                if len(context.pages) > pages_count_before:
                    print("检测到新开的账单标签页，正在切换...")
                    page = context.pages[-1]
                    page.bring_to_front()
                    page.wait_for_load_state("domcontentloaded", timeout=10000)
                    
            except Exception as e:
                print(f"⚠️ 导航到收支账单出现异常: {e}")
                
            page.get_by_text("支出账单", exact=True).first.click(force=True)
            page.wait_for_load_state("domcontentloaded", timeout=10000)
            page.wait_for_timeout(2000)
            
            # --- 填写业务大类 ---
            print("尝试点击下拉框并输入...")
            try:
                page.get_by_text("全部", exact=True).first.click(force=True)
                page.wait_for_timeout(1000)
            except:
                try:
                    page.get_by_text("请选择", exact=True).first.click(force=True)
                    page.wait_for_timeout(1000)
                except:
                    page.locator("input[readonly]").nth(1).click(force=True)
                    page.wait_for_timeout(1000)

            # 模拟键盘盲打输入类目名称
            page.keyboard.type(category, delay=100) 
            page.wait_for_timeout(2000) # 多等一会儿，确保天猫后端的联想接口返回结果
            
            # 【核心修改】：严格验证机制
            # 我们不去找“无业务大类选项”这种不确定的提示词了
            # 我们直接去找：下拉列表里，到底有没有出现我们要的这个类目！
            category_selected = False
            
            # 寻找下拉列表中包含我们类目名字的选项 (li 标签)
            target_option = page.locator(f"li[role='option']:has-text('{category}')").first
            
            if target_option.is_visible():
                # 如果找到了，强制点击它！这比敲回车更靠谱
                target_option.click(force=True)
                print(f"✅ 成功选中类目: {category}")
                category_selected = True
            else:
                # 如果没找到（不管是提示“无数据”还是什么别的原因），一律视为无选项
                # 尝试敲击回车作为最后挣扎
                page.keyboard.press("Enter")
                page.wait_for_timeout(1000)
                
                # 回车后，检查输入框里的值是不是我们要的（天猫的下拉框选中后，文字会显示在 span.next-select-inner 里）
                selected_text_locator = page.locator("span.next-select-inner:has-text('{}')".format(category)).first
                if selected_text_locator.is_visible():
                    print(f"✅ 虽未点到列表，但回车确认选中了: {category}")
                    category_selected = True
            
            # 【强制拦截关卡】
            if not category_selected:
                print(f"⚠️ 警告: 无法在下拉框中选中【{category}】（可能当月无数据或名称不匹配）。")
                print("🚫 直接跳过该大类的后续所有操作，进入下一个！")
                
                # 按一下 Esc 键收起可能卡在屏幕上的下拉菜单
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
                 # 【新增】：跳过前，清理掉本次循环新开的所有多余标签页
                while len(context.pages) > base_pages_count:
                    context.pages[-1].close()
                
                continue
                
            page.wait_for_timeout(1000)
            
            # --- 点击查询 ---
            try:
                page.get_by_role("button", name="查询").first.click(force=True)
            except:
                page.get_by_text("查询", exact=True).first.click(force=True)
            page.wait_for_timeout(3000) 

            # --- 抓取金额 ---
            amount = "0.00"
            try:
                row_locator = page.locator(f"tr:has-text('{category}'), div.row:has-text('{category}')").last
                row_text = row_locator.inner_text()
                # 切分文本提取包含小数点的金额数字
                text_parts = [t.strip() for t in row_text.split('\n') if t.strip()]
                for part in text_parts:
                    if "." in part and part.replace('.', '').replace(',', '').isdigit():
                        amount = part.replace(',', '').strip()
                        break
                print(f"✅ 成功提取金额: {amount}")
            except Exception as e:
                print(f"⚠️ 金额提取失败: {e}，将使用默认值 0.00")

            # --- 账单展开与截图 ---
            print("正在展开账单并截图...")
            try:
                # 寻找行首的展开按钮并点击
                expand_btn = page.locator(f"tr:has-text('{category}')").locator("span[class*='expand'], i[class*='expand']").first
                if expand_btn.is_visible():
                    expand_btn.click(force=True)
                    print("✅ 已点击账单明细展开按钮")
                    page.wait_for_timeout(1500) # 等待明细展开动画
                
                # 【核心】：注入 CSS，强制取消天猫表格的内部滚动条和高度限制
                page.evaluate("""
                    const style = document.createElement('style');
                    style.innerHTML = `
                        .next-table-body, .next-table-wrapper, 
                        .wn-table-body, .wn-table-wrapper {
                            max-height: none !important;
                            height: auto !important;
                            overflow: visible !important;
                        }
                    `;
                    document.head.appendChild(style);
                """)
                page.wait_for_timeout(500) # 等待样式生效，表格会被完全撑长
                
                # 此时内部滚动条已消失，表格被完全撑开，再执行全屏长图截图
                page.screenshot(path=images_paths['bill'], full_page=True)
                print("✅ 账单全屏长图截图完成")
            except Exception as e:
                print(f"⚠️ 账单展开失败，执行全屏长图截图: {e}")
                page.screenshot(path=images_paths['bill'], full_page=True)


            # ================= 第二阶段: 下载表格文件 =================
            print("点击下载全量明细，生成需要时间...")
            try:
                page.get_by_text("下载全量明细", exact=True).first.click(force=True)
            except:
                page.locator(f"tr:has-text('{category}')").locator("a:has-text('下载')").first.click(force=True)
            
            page.wait_for_timeout(2000) 
            
            # --- 处理各种可能出现的弹窗 ---
            try:
                # 寻找包含“已生成完成”或“已生成下载任务”的弹窗
                dialog_box = page.locator("div[role='dialog'], .next-dialog").filter(has_text=re.compile(r'已生成完成|已生成下载任务')).first
                
                if dialog_box.is_visible():
                    # 检查到底是哪种弹窗
                    dialog_text = dialog_box.inner_text()
                    
                    if "已生成下载任务" in dialog_text:
                        print("⚠️ 检测到'已生成下载任务'弹窗，正在等待 3 分钟让后台生成文件...")
                        page.wait_for_timeout(180000)  # 强制等待 180 秒 
                        print("3 分钟等待结束，正在关闭弹窗...")
                    elif "已生成完成" in dialog_text:
                        print("检测到'已生成完成'弹窗，正在关闭...")
                    
                    # 无论哪种弹窗，等待完毕后都尝试关闭它
                    close_x = dialog_box.locator(".next-dialog-close, i.next-icon-close").first
                    if close_x.is_visible():
                        close_x.click(force=True)
                    else:
                        # 兜底：如果找不到右上角的X，尝试点击“取消”或“确定”按钮
                        cancel_btn = dialog_box.locator("button:has-text('取消'), button:has-text('确定')").first
                        if cancel_btn.is_visible():
                            cancel_btn.click(force=True)
                            
                    page.wait_for_timeout(1000) 
            except Exception as e:
                # 如果找不到弹窗或者处理失败，直接忽略不影响主流程
                pass
            
            page.wait_for_timeout(3000) # 给后台生成文件留出一点余量时间
            
            # 点击进入历史记录
            page.get_by_text("历史下载记录", exact=True).first.click(force=True)
            page.wait_for_timeout(2000) 
            
            # --- 翻页寻找下载按钮并自动识别后缀 ---
            download_found = False
            temp_excel_path = ""
            for page_num in range(20):
                print(f"正在历史记录第 {page_num + 1} 页寻找下载按钮...")
                page.wait_for_timeout(1500) 
                
                # 基于 F12 截图特征的精准定位
                category_div = page.locator(f"div.next-table-cell-wrapper:has-text('{category}')").first
                
                if category_div.is_visible():
                    target_row = category_div.locator("xpath=ancestor::tr[contains(@class, 'next-table-row')]").first
                    download_btn = target_row.locator("a:has-text('下载'), button:has-text('下载')").first
                    
                    if download_btn.is_visible():
                        print(f"✅ 在第 {page_num + 1} 页找到【{category}】的下载按钮！开始下载...")
                        try:
                            with page.expect_download(timeout=15000) as download_info:
                                download_btn.click(force=True)
                            
                            download = download_info.value
                            original_filename = download.suggested_filename
                            _, ext = os.path.splitext(original_filename)
                            if not ext: ext = '.csv' # 天猫很多数据默认是无后缀的 csv 文本
                                
                            temp_excel_path = f"temp_excel_{category}{ext}"
                            download.save_as(temp_excel_path)
                            print(f"✅ 文件下载成功: {temp_excel_path}")
                            download_found = True
                            break # 下载成功，跳出翻页循环
                        except Exception as e:
                            print(f"⚠️ 点击下载时出错: {e}")
                
                if not download_found:
                    # 没找到则点击下一页
                    next_page_btn = page.locator(".next-pagination-item.next-next, button:has-text('下一页'), a:has-text('下一页')").first
                    if next_page_btn.is_visible() and not next_page_btn.get_attribute("disabled"):
                        next_page_btn.click(force=True)
                        page.wait_for_timeout(2000) 
                    else:
                        break
            
            # 关闭历史记录弹窗
            try:
                page.locator(".next-dialog-close, .close-icon").first.click(force=True)
                page.wait_for_timeout(1000)
            except:
                pass

            if not download_found:
                print("⚠️ 未能下载到账单明细，跳过发票环节，处理下一个业务大类。")
                # 【新增】：跳过前，清理掉本次循环新开的所有多余标签页
                while len(context.pages) > base_pages_count:
                    context.pages[-1].close()
                    
                continue

            # ================= 第三阶段: 发票截图 =================
            print("准备跳转至发票管理页面...")
            
            # 【核心修改】：不再依赖旧页面，直接新建一个干净的标签页
            try:
                page = context.new_page()
                page.bring_to_front()
                page.goto("https://myseller.taobao.com/", timeout=15000)
                page.wait_for_timeout(3000)
            except Exception as e:
                print(f"⚠️ 新建标签页失败，尝试使用当前最新页面: {e}")
                if len(context.pages) > 0:
                    page = context.pages[-1]
                    page.bring_to_front()
                    page.goto("https://myseller.taobao.com/")
                    page.wait_for_timeout(3000)
            
            # 重新在新页面里点击侧边栏导航
            try:
                # 寻找左侧菜单的“财务”
                page.get_by_text("财务", exact=True).first.click(force=True)
                page.wait_for_timeout(1500)
                
                # 记录一下点击前的页面数量
                pages_count_before = len(context.pages)
                
                page.get_by_text("申请发票", exact=True).first.click(force=True)
                page.wait_for_timeout(3000) # 给新页面弹出留足时间
                
                # 处理天猫可能弹出的新标签页
                if len(context.pages) > pages_count_before:
                    print("检测到新开的发票管理标签页，正在切换...")
                    page = context.pages[-1] 
                    page.bring_to_front()
                    page.wait_for_load_state("domcontentloaded", timeout=10000)
                
            except Exception as e:
                print(f"⚠️ 导航到发票模块时出现异常: {e}")

            page.wait_for_timeout(2000)
            
            print("尝试点击 '已申请发票' 标签...")
            try:
                # 策略 1: 基于 class 和文字的模糊匹配
                tab_locator = page.locator("div[class*='InvoiceRequest--tabItem']:has-text('已申请发票')").first
                if tab_locator.is_visible():
                    tab_locator.click(force=True)
                else:
                    # 策略 2: 去掉 exact=True 的模糊匹配
                    page.get_by_text("已申请发票").first.click(force=True)
            except Exception as e:
                print(f"⚠️ 点击 '已申请发票' 标签出现异常，尝试继续执行... 错误: {e}")
                
            page.wait_for_timeout(2000)
            
            # --- 筛选发票 ---
            print("正在筛选发票条件...")
            try:
                # 1. 选择业务类型
                print(f"尝试展开业务类型下拉框并输入: {category}")
                
                try:
                    # 使用你验证过的有效代码逻辑：直接按索引点击箭头
                    # 既然 nth(1) 能点开业务类型，我们就直接用它
                    page.locator("i.next-icon-arrow-down").nth(1).click(force=True)
                except Exception as e:
                    print(f"点击业务类型箭头失败，尝试点击前一个: {e}")
                    # 备用：如果页面结构变了，尝试点第一个
                    page.locator("i.next-icon-arrow-down").nth(0).click(force=True)
                
                page.wait_for_timeout(1000) # 等待下拉菜单展开
                
                # 开始模拟人工打字
                print(f"开始输入: {category}")
                page.keyboard.type(category, delay=150)
                page.wait_for_timeout(1500) # 等待联想菜单弹出来
                
                # 敲击回车，选中联想出来的结果
                page.keyboard.press("Enter")
                page.wait_for_timeout(1000)

                # 兜底：如果敲击回车没生效，尝试直接点击弹出的选项
                try:
                    page.locator(f"li[role='option']:has-text('{category}')").first.click(force=True, timeout=2000)
                except:
                    pass

                # 2. 选择申请时间为“最近一月”
                print("尝试选择申请时间: 最近一月")
                try:
                    # 既然业务类型是 nth(1)，那紧挨着它的申请时间必然是 nth(2)
                    page.locator("i.next-icon-arrow-down").nth(2).click(force=True)
                except:
                    # 备用方案，尝试用原有的备用逻辑
                    page.locator("i.next-icon-arrow-down").last.click(force=True)
                    
                page.wait_for_timeout(1000) 
                
                # 在弹出的列表中点击“最近一月”
                page.locator("li[role='option']:has-text('最近一月')").first.click(force=True)
                page.wait_for_timeout(1000)
                
                # 3. 点击搜索按钮
                print("点击搜索发票...")
                try:
                    page.get_by_role("button", name="搜索").first.click(force=True)
                except:
                    page.get_by_text("搜索", exact=True).first.click(force=True)
                    
                page.wait_for_timeout(3000) # 等待搜索结果加载
                print("✅ 发票筛选完成")
                
            except Exception as e:
                print(f"⚠️ 发票搜索条件输入异常，将直接对当前列表截图: {e}")

            # --- 发票列表展开并截图1 ---
            print("正在展开发票列表并获取全屏长图截图...")
            try:
                # 终极定位展开按钮：寻找表格第一列里那个看起来像按钮的图标
                # 天猫常用类名包括 next-table-expand-icon, fold-icon, plus-icon 等
                invoice_expand_btn = page.locator("td.first, td:first-child").locator("span[class*='expand'], i[class*='expand'], i[class*='fold']").first
                
                if invoice_expand_btn.is_visible():
                    # 点击它展开明细
                    invoice_expand_btn.click(force=True)
                    print("✅ 已点击发票列表展开按钮")
                    page.wait_for_timeout(1500) # 等待明细加载
                else:
                    # 如果上面定位不到，尝试直接寻找页面上第一个加号图标
                    fallback_expand = page.locator("i.next-icon-add, i.wn-icon-plus").first
                    if fallback_expand.is_visible():
                        fallback_expand.click(force=True)
                        print("✅ 尝试点击了备用的加号展开按钮")
                        page.wait_for_timeout(1500)
                
                # 【核心修改】：截取完整长图
                page.screenshot(path=images_paths['invoice1'], full_page=True)
                print("✅ 发票列表全屏长图截图1完成")
                    
            except Exception as e:
                print(f"⚠️ 发票列表展开失败，执行全屏长图截图: {e}")
                page.screenshot(path=images_paths['invoice1'], full_page=True)

            # --- 查看详情并截图2 ---
            print("正在获取发票详情全屏截图...")
            try:
                page.get_by_text("查看详情", exact=True).first.click(force=True)
                page.wait_for_timeout(2000)
                
                if len(context.pages) > 1:
                    detail_page = context.pages[-1]
                    detail_page.bring_to_front()
                    detail_page.wait_for_load_state("domcontentloaded", timeout=10000)
                    detail_page.wait_for_timeout(2000)
                    
                    # 【核心修改】：新标签页直接全屏截图
                    detail_page.screenshot(path=images_paths['invoice2'])
                    detail_page.close() 
                else:
                    # 【核心修改】：如果是弹窗，也直接截当前页面的全屏
                    page.screenshot(path=images_paths['invoice2'])
                    print("✅ 发票详情全屏截图2完成")
            except Exception as e:
                print(f"⚠️ 发票详情截图失败: {e}")


            # ================= 第四阶段: 数据整合 =================
            print("正在整合 Excel 和 所有截图文件...")
            process_excel_and_images(temp_excel_path, amount, category, images_paths)
            
            # 清理当前大类产生的临时图片和表格文件
            for path in images_paths.values():
                if os.path.exists(path):
                    os.remove(path)
            if os.path.exists(temp_excel_path):
                os.remove(temp_excel_path)

            # 当前大类全部处理完毕，清理掉本次循环产生的所有新标签页
            print(f"🧹 清理【{category}】相关的浏览器标签页...")
            while len(context.pages) > base_pages_count:
                context.pages[-1].close()
            
        print("\n🎉 所有流程执行完毕！")

if __name__ == "__main__":
    main()